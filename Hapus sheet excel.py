import os
from openpyxl import load_workbook
import shutil
import zipfile

# Fungsi untuk menghapus sheet "metadata kegiatan" jika ada
def delete_metadata_kegiatan_sheet(file_path):
    try:
        # Load workbook
        workbook = load_workbook(filename=file_path)
        
        # Cek apakah sheet "metadata kegiatan" ada
        if "Metadata Kegiatan" in workbook.sheetnames:
            # Jika sheet "metadata kegiatan" disembunyikan, ubah state-nya ke 'visible'
            if workbook["Metadata Kegiatan"].sheet_state == 'hidden':
                workbook["Metadata Kegiatan"].sheet_state = 'visible'
                
            # Hapus sheet "metadata kegiatan"
            workbook.remove(workbook["Metadata Kegiatan"])
            # Simpan perubahan
            workbook.save(file_path)
            print(f"Sheet 'metadata kegiatan' dihapus dari file {file_path}")
        else:
            print(f"Tidak ada sheet 'metadata kegiatan' di file {file_path}")
    except Exception as e:
        print(f"Gagal memproses file {file_path}: {e}")

# Path ke folder berisi file Excel
folder_path = "D:\Magang\Hapus metadata kegiatan\Badan Kepegawaian dan Pengembangan Sumber Daya Manusia\MS-Var"
output_folder = "Badan Kepegawaian dan Pengembangan Sumberdaya Manusia"


# Buat folder untuk menyimpan file yang sudah diproses
os.makedirs(output_folder, exist_ok=True)

# Loop untuk semua file di folder
for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx"):
        original_file_path = os.path.join(folder_path, file_name)
        processed_file_path = os.path.join(output_folder, file_name)
        
        # Salin file ke folder output
        shutil.copy2(original_file_path, processed_file_path)
        
        # Proses file di folder output
        delete_metadata_kegiatan_sheet(processed_file_path)

# Buat file ZIP dari folder output
zip_file_path = "processed_files.zip"
with zipfile.ZipFile(zip_file_path, 'w') as zipf:
    for root, _, files in os.walk(output_folder):
        for file in files:
            file_path = os.path.join(root, file)
            zipf.write(file_path, arcname=os.path.relpath(file_path, output_folder))

print(f"File ZIP siap diunduh: {zip_file_path}")